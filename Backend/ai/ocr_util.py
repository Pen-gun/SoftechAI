from paddleocr import PaddleOCR
from openpyxl import load_workbook
import os
import pandas as pd

# PaddleOCR expects a single language code; use 'en' (latin). For Nepali, switch to 'devanagari'.
ocr = PaddleOCR(lang='en', use_angle_cls=True)

def extract_text_from_file(file_path):
    """
    Extract text from PDF/image using PaddleOCR.
    Returns text as a string.
    """
    try:
        results = ocr.ocr(file_path, cls=True)
        if results is None or len(results) == 0:
            return "No text detected in the document."
        
        text = ""
        for line in results:
            if line is None:
                continue
            for word_info in line:
                text += word_info[1][0] + " "
            text += "\n"
        return text.strip() if text else "No text extracted."
    except Exception as e:
        raise ValueError(f"Error during OCR: {str(e)}")

def extract_text_from_excel(file_path):
    """
    Extract text from Excel file (.xlsx, .xls).
    Returns all cell values as a single string.
    """
    try:
        workbook = load_workbook(file_path)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + " "
                text += "\n"
            text += "\n"
        return text.strip()
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")

def answer_question_on_excel(file_path, question):
    """
    Answer questions on Excel data using pandas and keyword matching.
    Returns relevant rows/data matching the question.
    """
    try:
        # Read all sheets and search across them
        excel_file = pd.ExcelFile(file_path)
        all_results = []
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Convert question to lowercase for case-insensitive matching
            question_lower = question.lower()
            words = [w for w in question_lower.split() if len(w) > 2]  # Filter short words
            
            # Search across all cells
            for col in df.columns:
                try:
                    # Convert to string for searching
                    mask = df[col].astype(str).str.lower().str.contains('|'.join(words), na=False, regex=True)
                    matching = df[mask]
                    if not matching.empty:
                        all_results.append(f"Sheet '{sheet_name}', Column '{col}':\n{matching.to_string()}")
                except:
                    continue
        
        # Return results
        if all_results:
            answer = f"Question: {question}\n\nResults:\n" + "\n---\n".join(all_results[:3])  # Limit to top 3
        else:
            # If no matches, return data preview
            df = pd.read_excel(file_path, sheet_name=0)
            answer = f"Question: {question}\n\nNo exact match. Data preview:\n{df.head().to_string()}"
        
        return answer
    except Exception as e:
        return f"Error processing Excel: {str(e)}"
